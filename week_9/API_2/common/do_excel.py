# -*- coding: utf-8 -*-
# @Time    : 2019/3/11 20:31
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
from openpyxl import load_workbook
from week_9.API_2.common import project_path
from week_9.API_2.common.read_config import ReadConfig

class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel工作簿文件名或地址
        self.sheet_name=sheet_name#表单名

    def read_data(self):
        '''从Excel读取数据，有返回值'''
        case_id=ReadConfig(project_path.conf_path).get_data('CASE','case_id')
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        #唯一的要求是什么？每一行数据要在一起 {} []
        #如何把每一行的数据存到一个空间里面去？ []
        #开始读取数据
        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            row_data['Params']=sheet.cell(i,6).value#字符串 find ‘tel’ replace tel表单里面的电话号码
            row_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(row_data)
        wb.close()
        final_data=[]#空列表 存储最终的测试用例数据
        if case_id=='all':#如果case_id==all 那就获取所有的用例数据
           final_data=test_data#把测试用例赋值给final_data这个变量
        else:#否则 如果是列表 那就获取列表里面指定id的用例的数据
           for i in case_id:#遍历case_id 里面的值
               final_data.append(test_data[i-1])#？对应关系？？
        return final_data

    def write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        sheet.cell(row,col).value=value

        wb.save(self.file_name)
        wb.close()#关闭文件的动作

if __name__ == '__main__':
    file_name='test_api.xlsx'
    sheet_name='test_cases'
    test_data=DoExcel(project_path.case_path,sheet_name).read_data()
    print(test_data)
