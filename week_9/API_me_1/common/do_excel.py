#!/usr/bin/python
# -*- coding: <utf-8> -*-
#_Author_ = 'shuai'  
# Data：2019/12/28 10:26
from openpyxl import  load_workbook
class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name # Excel工作簿文件名或地址
        self.sheet_name = sheet_name #表单名
    def read_data(self):
        '''从Excel读取数据，有返回值'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        # 唯一的要求是什么？每一行数据要在一起 {} []
        # 如何把每一行的数据存到一个空间里面去？ []
        # 开始读取数据
        test_data = []
        for i in range(2, sheet.max_row + 1):#取头不取尾
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            row_data['Params'] = sheet.cell(i, 6).value
            row_data['ExpectedResult'] = sheet.cell(i, 7).value
            test_data.append(row_data)
        wb.close()
        return test_data

    def write_data(self, row, col, value):
        '''写回测试结果到Excel中'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    file_name = '../test_cases/test_api.xlsx'
    sheet_name = 'test_cases'
    test_data = DoExcel(file_name, sheet_name).read_data()
    print(test_data)
    # 结果是:
# [{'CaseId': 1, 'Module': 'Register', 'Title': '正常注册', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/register', 'Method': 'get', 'Params': "{'mobilephone':'18688775657','pwd':'123456'}", 'ExpectedResult': "{'status':1,'code':'10001','data': None,'msg':'注册成功'}"},
# {'CaseId': 2, 'Module': 'Register', 'Title': '不输入手机号', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/register', 'Method': 'post', 'Params': "{'mobilephone':'','pwd':'123456'}", 'ExpectedResult': "{'status':0,'code':'20103','data': None,'msg':'手机号不能为空'}"},
# {'CaseId': 3, 'Module': 'Register', 'Title': '不输入密码', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/register', 'Method': 'get', 'Params': "{'mobilephone':'18688775656','pwd':''}", 'ExpectedResult': "{'status':0,'code':'20103','data': None,'msg':'密码不能为空'}"},
# {'CaseId': 4, 'Module': 'Register', 'Title': '输入已存在的手机号', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/register', 'Method': 'post', 'Params': "{'mobilephone':'18688775656','pwd':'123456'}", 'ExpectedResult': "{'status':0,'code':'20110','data': None,'msg':'手机号码已被注册'}"},
# {'CaseId': 5, 'Module': 'Register', 'Title': '输入错误的手机号', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/register', 'Method': 'get', 'Params': "{'mobilephone':'186887756569','pwd':'123456'}", 'ExpectedResult': "{'status':0,'code':'20109','data': None,'msg':'手机号码格式不正确'}"},
# {'CaseId': 6, 'Module': 'Login', 'Title': '正常登录', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/login', 'Method': 'get', 'Params': "{'mobilephone':'18688775656','pwd':'123456'}", 'ExpectedResult': "{'status':1,'code':'10001','data': None,'msg':'登录成功'}"},
# {'CaseId': 7, 'Module': 'Login', 'Title': '不输入手机号', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/login', 'Method': 'post', 'Params': "{'mobilephone':'','pwd':'123456'}", 'ExpectedResult': "{'status':0,'code':'20103','data': None,'msg':'手机号不能为空'}"},
# {'CaseId': 8, 'Module': 'Login', 'Title': '不输入密码', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/login', 'Method': 'get', 'Params': "{'mobilephone':'18688775656','pwd':''}", 'ExpectedResult': "{'status':0,'code':'20103','data': None,'msg':'密码不能为空'}"},
# {'CaseId': 9, 'Module': 'Login', 'Title': '输入错误的手机号', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/login', 'Method': 'post', 'Params': "{'mobilephone':'186887756569','pwd':'123456'}", 'ExpectedResult': "{'status':0,'code':'20111','data': None,'msg':'用户名或密码错误'}"},
# {'CaseId': 10, 'Module': 'Login', 'Title': '输入错误的密码', 'Url': 'http://47.107.168.87:8080/futureloan/mvc/api/member/login', 'Method': 'get', 'Params': "{'mobilephone':'18688775656','pwd':'1234567'}", 'ExpectedResult': "{'status':0,'code':'20111','data': None,'msg':'用户名或密码错误'}"}]